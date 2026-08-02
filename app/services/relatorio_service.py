"""
app/services/relatorio_service.py

Gera relatórios de Premiles em dois formatos:
  - Excel (.xlsx) via openpyxl
  - PDF       via reportlab
"""
import io
from datetime import datetime

# ── Excel ─────────────────────────────────────────────────────────────────────

def gerar_excel(pgms, titulo="Relatório de Premiles"):
    """
    Recebe uma lista de objetos PGM (com .juniors e .juniors[n].balance).
    Retorna um BytesIO pronto para ser servido pelo Flask.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Premiles"

    # Paleta de cores TeoKids
    PRETO   = "FF0A0A0A"
    ROXO    = "FF7C3AED"
    CIANO   = "FF06B6D4"
    VERDE   = "FF10B981"
    LARANJA = "FFF97316"
    BRANCO  = "FFFFFFFF"
    CINZA   = "FF888888"
    BG_CARD = "FF141414"
    BG_ROW  = "FF1C1C1C"

    thin = Side(style="thin", color="FF2A2A2A")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = titulo
    c.font = Font(name="Calibri", bold=True, size=16, color=BRANCO)
    c.fill = PatternFill("solid", fgColor=ROXO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    c.font = Font(name="Calibri", size=10, color=CINZA)
    c.fill = PatternFill("solid", fgColor=BG_CARD)
    c.alignment = Alignment(horizontal="center")

    linha = 4

    for pgm in pgms:
        # Sub-título por PGM
        ws.merge_cells(f"A{linha}:E{linha}")
        c = ws.cell(row=linha, column=1, value=f"📋 {pgm.name}")
        c.font = Font(name="Calibri", bold=True, size=12, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=ROXO)
        c.alignment = Alignment(horizontal="left", indent=1)
        ws.row_dimensions[linha].height = 24
        linha += 1

        # Cabeçalho da tabela
        headers = ["#", "Nome", "PGM", "Premiles", "Posição"]
        cores_header = [BG_CARD, BG_CARD, BG_CARD, BG_CARD, BG_CARD]
        for col, (h, cor) in enumerate(zip(headers, cores_header), 1):
            c = ws.cell(row=linha, column=col, value=h)
            c.font = Font(name="Calibri", bold=True, size=9,
                          color=CINZA)
            c.fill = PatternFill("solid", fgColor=BG_CARD)
            c.alignment = Alignment(horizontal="center")
            c.border = borda
        ws.row_dimensions[linha].height = 18
        linha += 1

        # Juniores ordenados por saldo
        juniors_sorted = sorted(
            pgm.juniors,
            key=lambda j: j.balance.total_balance if j.balance else 0,
            reverse=True
        )

        for pos, junior in enumerate(juniors_sorted, 1):
            saldo = junior.balance.total_balance if junior.balance else 0
            medalha = {1: "🥇", 2: "🥈", 3: "🥉"}.get(pos, str(pos))
            fill = BG_ROW if pos % 2 == 0 else BG_CARD
            cor_pts = LARANJA if saldo > 0 else CINZA

            dados = [pos, junior.name, pgm.name, saldo, medalha]
            for col, val in enumerate(dados, 1):
                c = ws.cell(row=linha, column=col, value=val)
                c.font = Font(
                    name="Calibri", size=10,
                    color=cor_pts if col == 4 else BRANCO,
                    bold=(col == 4)
                )
                c.fill = PatternFill("solid", fgColor=fill)
                c.alignment = Alignment(horizontal="center" if col != 2 else "left",
                                        indent=1 if col == 2 else 0)
                c.border = borda
            ws.row_dimensions[linha].height = 20
            linha += 1

        # Linha de total do PGM
        total_pgm = sum(
            (j.balance.total_balance if j.balance else 0) for j in pgm.juniors
        )
        ws.merge_cells(f"A{linha}:C{linha}")
        c = ws.cell(row=linha, column=1, value=f"Total {pgm.name}")
        c.font = Font(name="Calibri", bold=True, size=10, color=BRANCO)
        c.fill = PatternFill("solid", fgColor="FF2A2A2A")
        c.alignment = Alignment(horizontal="right", indent=1)
        c.border = borda

        c = ws.cell(row=linha, column=4, value=total_pgm)
        c.font = Font(name="Calibri", bold=True, size=11, color=LARANJA)
        c.fill = PatternFill("solid", fgColor="FF2A2A2A")
        c.alignment = Alignment(horizontal="center")
        c.border = borda

        ws.cell(row=linha, column=5).fill = PatternFill("solid", fgColor="FF2A2A2A")
        ws.cell(row=linha, column=5).border = borda
        ws.row_dimensions[linha].height = 22
        linha += 2

    # Largura das colunas
    larguras = [6, 28, 20, 12, 10]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = larg

    # Congela cabeçalho
    ws.freeze_panes = "A4"

    # Fundo geral
    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── PDF ───────────────────────────────────────────────────────────────────────

def gerar_pdf(pgms, titulo="Relatório de Premiles"):
    """
    Recebe uma lista de objetos PGM.
    Retorna um BytesIO com o PDF gerado.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
        title=titulo
    )

    ROXO    = colors.HexColor("#7C3AED")
    CIANO   = colors.HexColor("#06B6D4")
    VERDE   = colors.HexColor("#10B981")
    LARANJA = colors.HexColor("#F97316")
    PRETO   = colors.HexColor("#0A0A0A")
    BRANCO  = colors.white
    CINZA   = colors.HexColor("#888888")
    BG_CARD = colors.HexColor("#141414")
    BG_ROW  = colors.HexColor("#1C1C1C")

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "Titulo", parent=styles["Title"],
        fontSize=20, textColor=BRANCO, spaceAfter=4,
        fontName="Helvetica-Bold", alignment=TA_CENTER,
        backColor=ROXO
    )
    estilo_sub = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        fontSize=9, textColor=CINZA, spaceAfter=16,
        fontName="Helvetica", alignment=TA_CENTER
    )
    estilo_pgm = ParagraphStyle(
        "PGM", parent=styles["Normal"],
        fontSize=13, textColor=BRANCO, spaceBefore=16, spaceAfter=6,
        fontName="Helvetica-Bold", backColor=ROXO, leftIndent=6
    )

    conteudo = []

    # Título
    conteudo.append(Paragraph(titulo, estilo_titulo))
    conteudo.append(Paragraph(
        f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
        estilo_sub
    ))

    for pgm in pgms:
        conteudo.append(Paragraph(f"  {pgm.name}", estilo_pgm))
        conteudo.append(Spacer(1, 4))

        juniors_sorted = sorted(
            pgm.juniors,
            key=lambda j: j.balance.total_balance if j.balance else 0,
            reverse=True
        )

        # Dados da tabela
        dados = [["#", "Nome", "Premiles"]]
        for pos, junior in enumerate(juniors_sorted, 1):
            saldo = junior.balance.total_balance if junior.balance else 0
            medalha = {1: "1o", 2: "2o", 3: "3o"}.get(pos, str(pos))
            dados.append([medalha, junior.name, str(saldo)])

        # Linha de total
        total_pgm = sum(
            (j.balance.total_balance if j.balance else 0) for j in pgm.juniors
        )
        dados.append(["", "TOTAL", str(total_pgm)])

        tabela = Table(dados, colWidths=[1.5*cm, 10*cm, 4*cm])
        tabela.setStyle(TableStyle([
            # Cabeçalho
            ("BACKGROUND",   (0, 0), (-1, 0),  ROXO),
            ("TEXTCOLOR",    (0, 0), (-1, 0),  BRANCO),
            ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, 0),  9),
            ("ALIGN",        (0, 0), (-1, 0),  "CENTER"),
            ("BOTTOMPADDING",(0, 0), (-1, 0),  8),
            ("TOPPADDING",   (0, 0), (-1, 0),  8),
            # Corpo
            ("BACKGROUND",   (0, 1), (-1, -2), BG_CARD),
            ("ROWBACKGROUNDS",(0,1), (-1,-2), [BG_CARD, BG_ROW]),
            ("TEXTCOLOR",    (0, 1), (-1, -2), BRANCO),
            ("FONTNAME",     (0, 1), (-1, -2), "Helvetica"),
            ("FONTSIZE",     (0, 1), (-1, -2), 9),
            ("ALIGN",        (0, 1), (0, -1),  "CENTER"),
            ("ALIGN",        (2, 1), (2, -1),  "CENTER"),
            ("BOTTOMPADDING",(0, 1), (-1, -2), 6),
            ("TOPPADDING",   (0, 1), (-1, -2), 6),
            # Coluna Premiles em laranja
            ("TEXTCOLOR",    (2, 1), (2, -2),  LARANJA),
            ("FONTNAME",     (2, 1), (2, -2),  "Helvetica-Bold"),
            # Linha de total
            ("BACKGROUND",   (0, -1), (-1, -1), colors.HexColor("#2A2A2A")),
            ("TEXTCOLOR",    (0, -1), (-1, -1), LARANJA),
            ("FONTNAME",     (0, -1), (-1, -1), "Helvetica-Bold"),
            ("ALIGN",        (1, -1), (1, -1),  "RIGHT"),
            ("TOPPADDING",   (0, -1), (-1, -1), 8),
            ("BOTTOMPADDING",(0, -1), (-1, -1), 8),
            # Grade
            ("GRID",         (0, 0), (-1, -1),  0.4, colors.HexColor("#2A2A2A")),
            ("ROUNDEDCORNERS", [4]),
        ]))

        conteudo.append(tabela)
        conteudo.append(Spacer(1, 8))

    # Rodapé
    conteudo.append(HRFlowable(width="100%", thickness=0.5, color=CINZA))
    conteudo.append(Paragraph(
        "TeoK!ds. — Projeto Juniores • Relatório gerado automaticamente",
        ParagraphStyle("Rodape", parent=styles["Normal"],
                       fontSize=7, textColor=CINZA, alignment=TA_CENTER,
                       spaceBefore=8)
    ))

    doc.build(conteudo)
    buf.seek(0)
    return buf
