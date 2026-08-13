from flask import render_template, redirect, url_for, flash, request
from flask_login import login_required, current_user
from app.services.audit_service import audit_log
from app.blueprints.supervisor import supervisor_bp
from app.decorators import requires_role
from app.models.user import UserRole, User
from app.models.pgm import PGM, PGMLeader
from app.models.activity import PremilesBalance
from app.models.audit import AuditLog, LogAction # ✅ Correto
from app import db


@supervisor_bp.route("/dashboard")
@login_required
@requires_role(UserRole.SUPERVISOR)
def dashboard():
    from app.services.period_service import get_active_period
    from app.models.activity import ChecklistLog, ManualLog
    from datetime import date as _date

    pgms = db.session.execute(db.select(PGM)).scalars().all()
    periodo_ativo = get_active_period()

    # Premiles creditados hoje
    hoje = _date.today()
    premiles_hoje = db.session.execute(
        db.select(db.func.sum(ManualLog.premiles_awarded))
        .where(db.func.date(ManualLog.created_at) == hoje)
    ).scalar() or 0
    premiles_hoje += db.session.execute(
        db.select(db.func.sum(ChecklistLog.premiles_awarded))
        .where(ChecklistLog.activity_date == hoje)
    ).scalar() or 0

    return render_template("supervisor/dashboard.html",
                           pgms=pgms,
                           periodo_ativo=periodo_ativo,
                           premiles_hoje=premiles_hoje)


@supervisor_bp.route("/pgm/<int:pgm_id>")
@login_required
@requires_role(UserRole.SUPERVISOR)
def pgm_detail(pgm_id):
    pgm = db.get_or_404(PGM, pgm_id)
    juniors = db.session.execute(
        db.select(User).where(User.pgm_id == pgm_id)
    ).scalars().all()
    return render_template("supervisor/pgm_detail.html", pgm=pgm, juniors=juniors)


@supervisor_bp.route("/junior/<int:junior_id>")
@login_required
@requires_role(UserRole.SUPERVISOR)
def junior_detail(junior_id):
    junior = db.get_or_404(User, junior_id)
    return render_template("supervisor/junior_detail.html", junior=junior)


# ── CRUD de Usuários ──────────────────────────────────────────────────────────

@supervisor_bp.route("/usuarios")
@login_required
@requires_role(UserRole.SUPERVISOR)
def usuarios():
    todos = db.session.execute(
        db.select(User).order_by(User.role, User.name)
    ).scalars().all()
    pgms = db.session.execute(db.select(PGM).order_by(PGM.name)).scalars().all()
    return render_template("supervisor/usuarios.html", usuarios=todos, pgms=pgms)


@supervisor_bp.route("/usuarios/novo", methods=["GET", "POST"])
@login_required
@requires_role(UserRole.SUPERVISOR)
def novo_usuario():
    pgms = db.session.execute(db.select(PGM).order_by(PGM.name)).scalars().all()

    if request.method == "POST":
        name     = request.form.get("name", "").strip()
        email    = request.form.get("email", "").strip().lower()
        role     = request.form.get("role")
        pgm_id   = request.form.get("pgm_id") or None
        password = request.form.get("password", "").strip()

        errors = []
        if not name:
            errors.append("Nome completo é obrigatório.")
        if not email:
            errors.append("E-mail é obrigatório.")
        if role not in [r.value for r in UserRole]:
            errors.append("Papel inválido.")
        if not password or len(password) < 6:
            errors.append("Senha deve ter ao menos 6 caracteres.")

        if email and db.session.execute(
            db.select(User).where(User.email == email)
        ).scalar_one_or_none():
            errors.append("Já existe um usuário com este e-mail.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("supervisor/form_usuario.html",
                                   pgms=pgms, modo="novo", form=request.form)

        user = User(
            name=name,
            email=email,
            role=UserRole(role),
            pgm_id=int(pgm_id) if pgm_id and role == UserRole.JUNIOR.value else None,
        )
        user.set_password(password)
        db.session.add(user)
        db.session.flush()

        if user.role == UserRole.JUNIOR:
            db.session.add(PremilesBalance(junior_id=user.id, total_balance=0))
        if user.role == UserRole.LIDER and pgm_id:
            db.session.add(PGMLeader(user_id=user.id, pgm_id=int(pgm_id)))

        audit_log(LogAction.USUARIO_CRIADO,
                  f"Usuário {name} ({email}) criado com papel {role}",
                  actor=current_user, target_user=user)
        db.session.commit()
        flash(f"Usuário {name} cadastrado com sucesso!", "success")
        return redirect(url_for("supervisor.usuarios"))

    return render_template("supervisor/form_usuario.html",
                           pgms=pgms, modo="novo", form={})


@supervisor_bp.route("/usuarios/<int:user_id>/editar", methods=["GET", "POST"])
@login_required
@requires_role(UserRole.SUPERVISOR)
def editar_usuario(user_id):
    user = db.get_or_404(User, user_id)
    pgms = db.session.execute(db.select(PGM).order_by(PGM.name)).scalars().all()

    if request.method == "POST":
        name       = request.form.get("name", "").strip()
        email      = request.form.get("email", "").strip().lower()
        pgm_id     = request.form.get("pgm_id") or None
        nova_senha = request.form.get("password", "").strip()

        errors = []
        if not name:
            errors.append("Nome completo é obrigatório.")
        if not email:
            errors.append("E-mail é obrigatório.")

        if email and db.session.execute(
            db.select(User).where(User.email == email, User.id != user_id)
        ).scalar_one_or_none():
            errors.append("Já existe outro usuário com este e-mail.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("supervisor/form_usuario.html",
                                   pgms=pgms, modo="editar",
                                   user=user, form=request.form)

        user.name  = name
        user.email = email

        # Atualiza PGM do júnior
        if user.role == UserRole.JUNIOR and pgm_id:
            user.pgm_id = int(pgm_id)

        # Atualiza vínculo do líder
        if user.role == UserRole.LIDER and pgm_id:
            vínculo = db.session.execute(
                db.select(PGMLeader).where(PGMLeader.user_id == user_id)
            ).scalar_one_or_none()
            if vínculo:
                vínculo.pgm_id = int(pgm_id)
            else:
                db.session.add(PGMLeader(user_id=user_id, pgm_id=int(pgm_id)))

        if nova_senha and len(nova_senha) >= 6:
            user.set_password(nova_senha)

        audit_log(LogAction.USUARIO_EDITADO,
                  f"Usuário {name} ({email}) editado",
                  actor=current_user, target_user=user)
        db.session.commit()
        flash(f"Usuário {name} atualizado com sucesso!", "success")
        return redirect(url_for("supervisor.usuarios"))

    return render_template("supervisor/form_usuario.html",
                           pgms=pgms, modo="editar",
                           user=user, form={})


@supervisor_bp.route("/usuarios/<int:user_id>/excluir", methods=["POST"])
@login_required
@requires_role(UserRole.SUPERVISOR)
def excluir_usuario(user_id):
    user = db.get_or_404(User, user_id)

    # Impede auto-exclusão
    if user.id == current_user.id:
        flash("Você não pode excluir sua própria conta.", "danger")
        return redirect(url_for("supervisor.usuarios"))

    nome = user.name
    audit_log(LogAction.USUARIO_EXCLUIDO,
              f"Usuário {nome} ({user.email}) foi excluído",
              actor=current_user)
    db.session.delete(user)
    db.session.commit()
    flash(f"Usuário {nome} removido.", "info")
    return redirect(url_for("supervisor.usuarios"))


@supervisor_bp.route("/usuarios/<int:user_id>/desbloquear", methods=["POST"])
@login_required
@requires_role(UserRole.SUPERVISOR)
def desbloquear_usuario(user_id):
    user = db.get_or_404(User, user_id)

    # Se a conta não estiver bloqueada, apenas avisa e volta
    if getattr(user, 'is_locked', False) is False:
        flash(f"A conta de {user.name} não está bloqueada.", "info")
        return redirect(url_for("supervisor.usuarios"))

    # Remove a trava e zera os erros
    user.is_locked = False
    user.failed_attempts = 0
    
    # Registra no log usando USUARIO_EDITADO para manter compatibilidade com o banco
    audit_log(
        LogAction.USUARIO_EDITADO,
        f"A conta de {user.name} ({user.email}) foi desbloqueada manualmente",
        actor=current_user, target_user=user
    )
    
    db.session.commit()
    flash(f"✅ Conta de {user.name} foi desbloqueada com sucesso!", "success")
    
    return redirect(url_for("supervisor.usuarios"))


# ── Ranking ───────────────────────────────────────────────────────────────────

@supervisor_bp.route("/ranking")
@login_required
@requires_role(UserRole.SUPERVISOR)
def ranking():
    pgms = db.session.execute(db.select(PGM)).scalars().all()

    # Todos os juniores ordenados por saldo desc
    juniors = db.session.execute(
        db.select(User)
        .where(User.role == UserRole.JUNIOR)
        .join(PremilesBalance, PremilesBalance.junior_id == User.id, isouter=True)
        .order_by(db.desc(PremilesBalance.total_balance))
    ).scalars().all()

    return render_template("supervisor/ranking.html", juniors=juniors, pgms=pgms)


# ── Exportação de relatórios ──────────────────────────────────────────────────

@supervisor_bp.route("/relatorio/excel")
@login_required
@requires_role(UserRole.SUPERVISOR)
def relatorio_excel():
    from flask import send_file
    from app.services.relatorio_service import gerar_excel
    from datetime import datetime

    pgms = db.session.execute(db.select(PGM)).scalars().all()
    buf = gerar_excel(pgms, titulo="Relatório de Premiles — Projeto Juniores")

    nome = f"premiles_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nome,
    )


@supervisor_bp.route("/relatorio/pdf")
@login_required
@requires_role(UserRole.SUPERVISOR)
def relatorio_pdf():
    from flask import send_file
    from app.services.relatorio_service import gerar_pdf
    from datetime import datetime

    pgms = db.session.execute(db.select(PGM)).scalars().all()
    buf = gerar_pdf(pgms, titulo="Relatório de Premiles — Projeto Juniores")

    nome = f"premiles_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=nome,
    )


@supervisor_bp.route("/relatorios")
@login_required
@requires_role(UserRole.SUPERVISOR)
def relatorios():
    pgms = db.session.execute(db.select(PGM)).scalars().all()
    return render_template("supervisor/relatorios.html", pgms=pgms)


# ── Recuperação de senha ──────────────────────────────────────────────────────

@supervisor_bp.route("/usuarios/<int:user_id>/gerar-link", methods=["POST"])
@login_required
@requires_role(UserRole.SUPERVISOR)
def gerar_link_reset(user_id):
    from flask import request as req
    user = db.get_or_404(User, user_id)
    token = user.generate_reset_token(hours=24)
    audit_log(LogAction.RESET_GERADO,
              f"Link de recuperação de senha gerado para {user.name} ({user.email})",
              actor=current_user, target_user=user)
    db.session.commit()
    link = req.host_url.rstrip("/") + f"/redefinir/{token}"
    flash(
        f"Link gerado para {user.name} (válido por 24h). "
        f"Copie e envie pelo WhatsApp: {link}",
        "info"
    )
    return redirect(url_for("supervisor.usuarios"))


# ── Configuração de atividades ────────────────────────────────────────────────

@supervisor_bp.route("/atividades")
@login_required
@requires_role(UserRole.SUPERVISOR)
def atividades():
    from app.models.activity import ActivityType, ActivitySource
    leader_acts = db.session.execute(
        db.select(ActivityType)
        .where(ActivityType.source == ActivitySource.LEADER)
        .order_by(ActivityType.name)
    ).scalars().all()
    junior_acts = db.session.execute(
        db.select(ActivityType)
        .where(ActivityType.source == ActivitySource.JUNIOR)
        .order_by(ActivityType.name)
    ).scalars().all()
    return render_template("supervisor/atividades.html",
                           leader_acts=leader_acts, junior_acts=junior_acts)


@supervisor_bp.route("/atividades/nova", methods=["POST"])
@login_required
@requires_role(UserRole.SUPERVISOR)
def nova_atividade():
    from app.models.activity import ActivityType, ActivitySource
    name             = request.form.get("name", "").strip()
    source           = request.form.get("source", "")
    default_premiles = int(request.form.get("default_premiles", 10))
    requires_summary = bool(request.form.get("requires_summary"))

    if not name:
        flash("Nome da atividade é obrigatório.", "danger")
        return redirect(url_for("supervisor.atividades"))
    if source not in ["leader", "junior"]:
        flash("Tipo inválido.", "danger")
        return redirect(url_for("supervisor.atividades"))

    exists = db.session.execute(
        db.select(ActivityType).where(ActivityType.name == name)
    ).scalar_one_or_none()
    if exists:
        flash(f'Já existe uma atividade chamada "{name}".', "warning")
        return redirect(url_for("supervisor.atividades"))

    new_act = ActivityType(
        name=name,
        source=ActivitySource(source),
        default_premiles=default_premiles,
        requires_summary=requires_summary,
    )
    db.session.add(new_act)
    audit_log(LogAction.ATIVIDADE_CRIADA,
              f'Atividade "{name}" criada ({source}, {default_premiles} pts)',
              actor=current_user)
    db.session.commit()
    flash(f'Atividade "{name}" criada com sucesso!', "success")
    return redirect(url_for("supervisor.atividades"))


@supervisor_bp.route("/atividades/<int:act_id>/editar", methods=["POST"])
@login_required
@requires_role(UserRole.SUPERVISOR)
def editar_atividade(act_id):
    from app.models.activity import ActivityType
    act = db.get_or_404(ActivityType, act_id)
    act.name             = request.form.get("name", act.name).strip()
    act.default_premiles = int(request.form.get("default_premiles", act.default_premiles))
    act.requires_summary = bool(request.form.get("requires_summary"))
    db.session.commit()
    flash(f'Atividade "{act.name}" atualizada.', "success")
    return redirect(url_for("supervisor.atividades"))


@supervisor_bp.route("/atividades/<int:act_id>/toggle", methods=["POST"])
@login_required
@requires_role(UserRole.SUPERVISOR)
def toggle_atividade(act_id):
    from app.models.activity import ActivityType
    act = db.get_or_404(ActivityType, act_id)
    act.is_active = not act.is_active
    estado = "ativada" if act.is_active else "desativada"
    audit_log(LogAction.ATIVIDADE_TOGGLE,
              f'Atividade "{act.name}" {estado}',
              actor=current_user)
    db.session.commit()
    flash(f'Atividade "{act.name}" {estado}.', "info")
    return redirect(url_for("supervisor.atividades"))


# ── Exportação do resultado de um período específico ──────────────────────────

@supervisor_bp.route("/periodos/<int:periodo_id>/exportar/excel")
@login_required
@requires_role(UserRole.SUPERVISOR)
def exportar_periodo_excel(periodo_id):
    from flask import send_file
    from app.models.period import Period, PeriodSnapshot
    from app.models.pgm import PGM
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    periodo = db.get_or_404(Period, periodo_id)
    snapshots = db.session.execute(
        db.select(PeriodSnapshot)
        .where(PeriodSnapshot.period_id == periodo_id)
        .order_by(PeriodSnapshot.position)
    ).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultado"
    ws.sheet_view.showGridLines = False

    ROXO  = "FF7C3AED"
    CINZA = "FF888888"
    BRANCO = "FFFFFFFF"
    LARANJA = "FFF97316"
    BG   = "FF141414"
    BG2  = "FF1C1C1C"
    thin = Side(style="thin", color="FF2A2A2A")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"Resultado — {periodo.name}"
    c.font = Font(name="Calibri", bold=True, size=14, color=BRANCO)
    c.fill = PatternFill("solid", fgColor=ROXO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = (f"{periodo.start_date.strftime('%d/%m/%Y')} → "
               f"{periodo.end_date.strftime('%d/%m/%Y') if periodo.end_date else 'em aberto'}"
               f" · Exportado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.font = Font(name="Calibri", size=9, color=CINZA)
    c.fill = PatternFill("solid", fgColor=BG)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    linha = 4
    headers = ["Pos.", "Nome", "PGM", "Premiles", "Pos. PGM"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=linha, column=col, value=h)
        c.font = Font(name="Calibri", bold=True, size=9, color=CINZA)
        c.fill = PatternFill("solid", fgColor=BG)
        c.alignment = Alignment(horizontal="center")
        c.border = borda
    ws.row_dimensions[linha].height = 18
    linha += 1

    for snap in snapshots:
        medalha = {1:"🥇",2:"🥈",3:"🥉"}.get(snap.position, f"{snap.position}º")
        fill = BG if snap.position % 2 == 1 else BG2
        dados = [medalha, snap.junior.name,
                 snap.pgm.name if snap.pgm else "—",
                 snap.total_premiles,
                 f"{snap.pgm_position}º" if snap.pgm_position else "—"]
        for col, val in enumerate(dados, 1):
            c = ws.cell(row=linha, column=col, value=val)
            c.font = Font(name="Calibri", size=10,
                          color=LARANJA if col == 4 else BRANCO,
                          bold=(col == 4))
            c.fill = PatternFill("solid", fgColor=fill)
            c.alignment = Alignment(horizontal="center" if col != 2 else "left", indent=1 if col==2 else 0)
            c.border = borda
        ws.row_dimensions[linha].height = 20
        linha += 1

    for col, larg in zip(range(1,6), [8, 26, 18, 12, 10]):
        ws.column_dimensions[get_column_letter(col)].width = larg

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"resultado_{periodo.name.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=nome)


@supervisor_bp.route("/periodos/<int:periodo_id>/exportar/pdf")
@login_required
@requires_role(UserRole.SUPERVISOR)
def exportar_periodo_pdf(periodo_id):
    from flask import send_file
    from app.models.period import Period, PeriodSnapshot
    from app.services.relatorio_service import gerar_pdf
    from datetime import datetime
    import io

    periodo = db.get_or_404(Period, periodo_id)
    snapshots = db.session.execute(
        db.select(PeriodSnapshot)
        .where(PeriodSnapshot.period_id == periodo_id)
        .order_by(PeriodSnapshot.position)
    ).scalars().all()

    # Agrupa snapshots por PGM simulando estrutura de pgms
    class FakePGM:
        def __init__(self, name, juniors):
            self.name = name
            self.juniors = juniors

    class FakeJunior:
        def __init__(self, snap):
            self.name = snap.junior.name
            self.balance = type("b", (), {"total_balance": snap.total_premiles})()

    pgm_map = {}
    for snap in snapshots:
        pgm_name = snap.pgm.name if snap.pgm else "Sem PGM"
        if pgm_name not in pgm_map:
            pgm_map[pgm_name] = []
        pgm_map[pgm_name].append(FakeJunior(snap))

    fake_pgms = [FakePGM(name, juniors) for name, juniors in pgm_map.items()]
    titulo = f"Resultado — {periodo.name}"
    buf = gerar_pdf(fake_pgms, titulo=titulo)

    nome = f"resultado_{periodo.name.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name=nome)


# ── Painel de logs de auditoria ───────────────────────────────────────────────

@supervisor_bp.route("/logs")
@login_required
@requires_role(UserRole.SUPERVISOR)
def logs():
    from app.models.audit import AuditLog, LogAction

    # Filtros
    user_id   = request.args.get("user_id", type=int)
    action    = request.args.get("action", "")
    page      = request.args.get("page", 1, type=int)
    per_page  = 50

    query = db.select(AuditLog).order_by(AuditLog.created_at.desc())

    if user_id:
        query = query.where(
            db.or_(
                AuditLog.actor_id == user_id,
                AuditLog.target_user_id == user_id
            )
        )
    if action:
        query = query.where(AuditLog.action == action)

    # Paginação manual
    total = db.session.execute(
        db.select(db.func.count()).select_from(query.subquery())
    ).scalar()

    logs_page = db.session.execute(
        query.offset((page - 1) * per_page).limit(per_page)
    ).scalars().all()

    # Para o filtro de usuários
    usuarios = db.session.execute(
        db.select(User).order_by(User.name)
    ).scalars().all()

    total_pages = (total + per_page - 1) // per_page

    return render_template(
        "supervisor/logs.html",
        logs=logs_page,
        usuarios=usuarios,
        actions=LogAction,
        filtro_user_id=user_id,
        filtro_action=action,
        page=page,
        total_pages=total_pages,
        total=total,
    )


# ── Zerar todos os saldos de Premiles e Extratos ─────────────────────────────

@supervisor_bp.route("/zerar-premiles", methods=["POST"])
@login_required
@requires_role(UserRole.SUPERVISOR)
def zerar_premiles():
    # Importamos os logs para poder apagá-los
    from app.models.activity import PremilesBalance, ManualLog, ChecklistLog
    
    confirmacao = request.form.get("confirmacao", "").strip()

    # Dupla confirmação: o supervisor precisa digitar "ZERAR" para confirmar
    if confirmacao != "ZERAR":
        flash("Confirmação incorreta. Digite ZERAR para confirmar a operação.", "danger")
        return redirect(url_for("supervisor.dashboard"))

    total_usuarios = db.session.execute(
        db.select(db.func.count(PremilesBalance.id))
    ).scalar()

    # 1. Zera os saldos de todos os juniores
    db.session.execute(
        db.update(PremilesBalance).values(total_balance=0)
    )

    # 2. Limpa os extratos (Apaga os históricos de lançamentos e checklists)
    db.session.execute(db.delete(ManualLog))
    db.session.execute(db.delete(ChecklistLog))

    # 3. Registra a auditoria
    audit_log(
        "LANCAMENTO_EXCLUIDO",
        f"Todos os saldos de Premiles foram zerados e os extratos foram limpos por {current_user.name} "
        f"({total_usuarios} juniores afetados)",
        actor=current_user
    )

    db.session.commit()
    
    flash(
        f"✅ Saldos e extratos zerados com sucesso! {total_usuarios} júnior(es) começaram um novo ciclo.",
        "success"
    )
    return redirect(url_for("supervisor.dashboard"))